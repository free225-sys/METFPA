"""Regenerate the Excel dry-run test fixtures used by metfpa_phase2a_test.py.

These workbooks live in /tmp/imp (ephemeral) and are recreated by running:
    python backend/tests/_gen_import_fixtures.py
Matches the approved SHEET_SPEC in backend/metfpa/imports.py.
"""
import os
from openpyxl import Workbook

OUT = "/tmp/imp"
os.makedirs(OUT, exist_ok=True)

# Required sheets with required columns + one valid row each (no warnings/errors).
VALID_SHEETS = {
    "PND": (["code", "parent_code", "niveau", "nom", "budget_total"],
            [["4", "", "pilier", "Capital humain", 1000000],
             ["4.02", "", "secteur", "EFTP", 1202137.6]]),
    "Politique_EFTP": (["code", "parent_code", "niveau", "nom", "budget_total"],
                       [["AX1", "", "axe", "Accès et équité", 500000]]),
    "Strategie_Digitale": (["code", "parent_code", "niveau", "nom", "budget_total"],
                           [["D1", "", "axe", "Infrastructure numérique", 33562]]),
    "Activites": (["code_action", "intitule", "axe_pol", "produit_pol", "direction", "statut", "avancement", "echeance", "budget_prevu"],
                  [["A001", "Construire un CFP", "AX1", "1.1", "DAF", "En cours", 40, "2026-T2", 1500],
                   ["A002", "Équiper un atelier", "AX1", "1.2", "DRH", "Non démarré", 0, "2026-T3", 800]]),
    "Indicateurs": (["niveau", "libelle", "base", "cible", "source"],
                    [["PND (national)", "Taux d'insertion", "n.d.", "60%", "Enquête"]]),
    "Alignements": (["pol_axe", "pnd_effet"], [["AX1", "4.02.1"]]),
    "Budget": (["framework", "period_start", "period_end", "total"],
               [["PND", 2026, 2030, 1202137.6]]),
    "Directions": (["code", "nom"], [["DAF", "Direction des Affaires Financières"]]),
}


def _write(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)


def _base_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in VALID_SHEETS.items():
        _write(wb, name, headers, [list(r) for r in rows])
    return wb


# 1) valid.xlsx -> READY_FOR_REVIEW
wb = _base_workbook()
wb.save(f"{OUT}/valid.xlsx")

# 2) warn.xlsx -> READY_WITH_WARNINGS (extra unknown sheet)
wb = _base_workbook()
_write(wb, "Feuille_Inconnue", ["x", "y"], [["a", "b"]])
wb.save(f"{OUT}/warn.xlsx")

# 3) blocked.xlsx -> BLOCKED_BY_ERRORS (invalid statut + negative budget)
wb = _base_workbook()
ws = wb["Activites"]
ws.append(["A003", "Activité fautive", "AX1", "1.3", "DAF", "STATUT_INVALIDE", 40, "2026-T1", -500])
wb.save(f"{OUT}/blocked.xlsx")

# 4) missingcol.xlsx -> BLOCKED_BY_ERRORS (Activites missing required 'statut' column)
wb = Workbook()
wb.remove(wb.active)
for name, (headers, rows) in VALID_SHEETS.items():
    if name == "Activites":
        h = [c for c in headers if c != "statut"]
        rows = [[v for c, v in zip(headers, r) if c != "statut"] for r in rows]
        _write(wb, name, h, rows)
    else:
        _write(wb, name, headers, [list(r) for r in rows])
wb.save(f"{OUT}/missingcol.xlsx")

# 5) fake.xls -> not a real xlsx (rejected by extension + ZIP signature)
with open(f"{OUT}/fake.xls", "wb") as f:
    f.write(b"This is not a real Excel workbook.")

print("Fixtures written to", OUT, ":", sorted(os.listdir(OUT)))
