"""Phase 2A-1 — Excel import DRY-RUN only (no business-data mutation, no apply).
Validates, classifies and compares proposed changes from IMPORT_METFPA.xlsx.
Writes only import-job metadata + validation results + audit. Temp files deleted."""
import io
import os
import uuid
import hashlib
import zipfile
from datetime import datetime, timezone

import openpyxl
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException

from .db import mdb, audit
from .auth import require_role

imports_router = APIRouter(prefix="/api/metfpa/imports")

MAX_BYTES = int(os.environ.get("METFPA_IMPORT_MAX_BYTES", str(10 * 1024 * 1024)))
ALLOWED_STATUTS = {"Non démarré", "À l'heure", "En cours", "En retard", "Bloqué", "Achevé", "Suspendu"}
IMPORT_ROLES = ("me_validator", "admin")

# Approved workbook specification (sheet -> rules). Used for this dry-run.
SHEET_SPEC = {
    "PND": {"required": ["code", "parent_code", "niveau", "nom"], "code": "code", "parent": "parent_code",
            "collection": "pnd_nodes", "match": "code", "budget": ["budget_total"]},
    "Politique_EFTP": {"required": ["code", "parent_code", "niveau", "nom"], "code": "code", "parent": "parent_code",
                       "collection": "pol_nodes", "match": "code", "budget": ["budget_total"]},
    "Strategie_Digitale": {"required": ["code", "parent_code", "niveau", "nom"], "code": "code", "parent": "parent_code",
                           "collection": "dig_nodes", "match": "code", "budget": ["budget_total"]},
    "Activites": {"required": ["code_action", "intitule", "axe_pol", "produit_pol", "direction", "statut"],
                  "code": "code_action", "collection": "activities", "match": "code_action",
                  "statut": "statut", "avancement": "avancement", "echeance": "echeance",
                  "budget": ["budget_prevu", "budget_engage", "budget_execute"]},
    "Indicateurs": {"required": ["niveau", "libelle", "cible"], "code": "libelle",
                    "collection": "indicators", "match": "libelle"},
    "Alignements": {"required": ["pol_axe", "pnd_effet"], "code": "pol_axe",
                    "collection": "alignments", "match": "pol_axe"},
    "Budget": {"required": ["framework", "period_start", "period_end", "total"], "code": "framework",
               "collection": "frameworks", "match": "key",
               "budget": ["total"], "periods": ["period_start", "period_end"]},
    "Directions": {"required": ["code", "nom"], "code": "code", "collection": "directions", "match": "code"},
}
REQUIRED_SHEETS = list(SHEET_SPEC.keys())


def _num(v):
    if v is None or v == "":
        return None, False
    try:
        return float(v), True
    except (ValueError, TypeError):
        return None, False


def _is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def _read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for i, r in enumerate(rows[1:], start=2):
        if all(c is None or c == "" for c in r):
            continue
        data.append((i, dict(zip(headers, r))))
    return headers, data


async def _existing_codes(collection, match):
    docs = await mdb[collection].find({}, {"_id": 0}).to_list(5000)
    out = {}
    for d in docs:
        key = d.get(match)
        if key is not None:
            out[str(key)] = d
    return out


async def _validate_workbook(filename, content, actor):
    report = {
        "import_id": uuid.uuid4().hex, "filename": filename,
        "sha256": hashlib.sha256(content).hexdigest(), "size_bytes": len(content),
        "uploader": actor, "processed_at": datetime.now(timezone.utc).isoformat(),
        "sheets": {}, "totals": {"rows": 0, "valid": 0, "warnings": 0, "errors": 0,
                                 "insert": 0, "update": 0, "unchanged": 0, "conflict": 0, "reject": 0},
        "file_errors": [], "rows": [], "verdict": None,
    }

    # File-level checks
    if not filename.lower().endswith(".xlsx"):
        report["file_errors"].append("Extension non autorisée (seul .xlsx est accepté).")
    if filename.lower().endswith((".xlsm", ".xls")):
        report["file_errors"].append("Format à macros ou ancien format non autorisé.")
    if len(content) > MAX_BYTES:
        report["file_errors"].append(f"Fichier trop volumineux (> {MAX_BYTES} octets).")
    if not content[:4] == b"PK\x03\x04":
        report["file_errors"].append("Archive invalide ou fichier corrompu (signature ZIP absente).")

    wb = None
    if not report["file_errors"]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_vba=False)
        except zipfile.BadZipFile:
            report["file_errors"].append("Classeur illisible (archive corrompue).")
        except Exception as e:
            msg = str(e).lower()
            if "encrypt" in msg or "password" in msg:
                report["file_errors"].append("Classeur protégé par mot de passe — non autorisé.")
            else:
                report["file_errors"].append("Classeur illisible.")

    if wb is not None:
        present = wb.sheetnames
        # macro-enabled detection
        if getattr(wb, "vba_archive", None):
            report["file_errors"].append("Classeur contenant des macros — non autorisé.")
        if not present:
            report["file_errors"].append("Classeur vide.")
        missing = [s for s in REQUIRED_SHEETS if s not in present]
        unknown = [s for s in present if s not in SHEET_SPEC]
        report["missing_sheets"] = missing
        report["unknown_sheets"] = unknown

        for sname in present:
            if sname not in SHEET_SPEC:
                report["sheets"][sname] = {"status": "unknown", "rows": 0,
                                           "note": "Feuille non reconnue — ignorée."}
                continue
            spec = SHEET_SPEC[sname]
            ws = wb[sname]
            headers, data = _read_sheet(ws)
            existing = await _existing_codes(spec["collection"], spec["match"])
            seen_codes = set()
            missing_cols = [c for c in spec["required"] if c not in headers]
            unknown_cols = [c for c in headers if c and c not in _allowed_cols(spec)]
            srep = {"status": "ok", "rows": len(data), "missing_columns": missing_cols,
                    "unknown_columns": unknown_cols, "insert": 0, "update": 0,
                    "unchanged": 0, "conflict": 0, "reject": 0}
            if missing_cols:
                srep["status"] = "schema_error"

            for rownum, row in data:
                errors, warnings = [], []
                for c in missing_cols:
                    errors.append(f"Colonne requise manquante : {c}")
                # formula cells
                for col, val in row.items():
                    if _is_formula(val):
                        warnings.append(f"Formule non évaluée en '{col}' (valeur ignorée).")
                code = row.get(spec["code"])
                code = str(code).strip() if code not in (None, "") else None
                if not code:
                    errors.append(f"Code/identifiant manquant ({spec['code']}).")
                elif code in seen_codes:
                    errors.append(f"Code dupliqué dans le fichier : {code}")
                if code:
                    seen_codes.add(code)

                # referential: parent must exist in-sheet
                if spec.get("parent"):
                    pc = row.get(spec["parent"])
                    pc = str(pc).strip() if pc not in (None, "") else None
                    if pc and pc not in {str(r[1].get(spec["code"])).strip() for r in data if r[1].get(spec["code"])}:
                        warnings.append(f"Référence parent introuvable dans la feuille : {pc}")

                # statut nomenclature
                if spec.get("statut") and row.get(spec["statut"]) not in (None, "") and row.get(spec["statut"]) not in ALLOWED_STATUTS:
                    errors.append(f"Statut hors nomenclature : {row.get(spec['statut'])}")
                # avancement 0-100
                if spec.get("avancement") and row.get(spec["avancement"]) not in (None, ""):
                    av, ok = _num(row.get(spec["avancement"]))
                    if not ok or av < 0 or av > 100:
                        errors.append(f"Avancement invalide (0–100 attendu) : {row.get(spec['avancement'])}")
                # echeance format YYYY-TQ
                if spec.get("echeance") and row.get(spec["echeance"]) not in (None, ""):
                    ech = str(row.get(spec["echeance"]))
                    if not (len(ech) == 7 and ech[4:6] == "-T" and ech[6] in "1234"):
                        warnings.append(f"Échéance non conforme (YYYY-TQ) : {ech}")
                # budget numeric/negative; missing preserved
                for bcol in spec.get("budget", []):
                    if bcol in row and row.get(bcol) not in (None, ""):
                        bv, ok = _num(row.get(bcol))
                        if not ok:
                            errors.append(f"Budget non numérique '{bcol}' : {row.get(bcol)}")
                        elif bv < 0:
                            errors.append(f"Budget négatif '{bcol}' : {bv}")
                # periods consistency
                if spec.get("periods"):
                    ps, ok1 = _num(row.get("period_start"))
                    pe, ok2 = _num(row.get("period_end"))
                    if ok1 and ok2 and ps > pe:
                        errors.append(f"Période incohérente : début {int(ps)} > fin {int(pe)}")

                # operation classification
                if errors:
                    op = "reject"
                elif code and code in existing:
                    cur = existing[code]
                    nom_field = "nom" if "nom" in row else ("libelle" if "libelle" in row else None)
                    differs = bool(nom_field and str(row.get(nom_field)) != str(cur.get(nom_field)))
                    op = "update" if differs else "unchanged"
                else:
                    op = "insert"

                srep[op] += 1
                report["totals"][op] += 1
                report["totals"]["rows"] += 1
                if op != "reject":
                    report["totals"]["valid"] += 1
                report["totals"]["warnings"] += len(warnings)
                report["totals"]["errors"] += len(errors)
                report["rows"].append({
                    "sheet": sname, "row": rownum, "code": code, "operation": op,
                    "errors": errors, "warnings": warnings,
                    "proposed_data_origin": "official_reference" if sname in ("PND", "Politique_EFTP", "Strategie_Digitale", "Budget", "Indicateurs", "Alignements", "Directions") else "demo_tracking",
                    "proposed_validation_status": "pending_metfpa_validation",
                })
            report["sheets"][sname] = srep
        wb.close()

    # verdict
    has_errors = bool(report["file_errors"]) or report["totals"]["errors"] > 0 or \
        any(s.get("status") == "schema_error" for s in report["sheets"].values()) or \
        bool(report.get("missing_sheets"))
    if has_errors:
        report["verdict"] = "BLOCKED_BY_ERRORS"
    elif report["totals"]["warnings"] > 0 or report.get("unknown_sheets"):
        report["verdict"] = "READY_WITH_WARNINGS"
    else:
        report["verdict"] = "READY_FOR_REVIEW"
    return report


def _allowed_cols(spec):
    cols = set(spec["required"])
    for k in ("parent", "code", "statut", "avancement", "echeance"):
        if spec.get(k):
            cols.add(spec[k])
    for b in spec.get("budget", []):
        cols.add(b)
    for p in spec.get("periods", []):
        cols.add(p)
    cols.update(["niveau", "nom", "libelle", "base", "cible", "source", "axe", "budget_scope",
                 "source_document", "pnd_effet", "pol_axe", "produit_pol", "axe_pol", "annual_average"])
    return cols


@imports_router.post("/excel/dry-run")
async def excel_dry_run(file: UploadFile = File(...), identity: dict = Depends(require_role(*IMPORT_ROLES))):
    content = await file.read()
    await audit("import_upload", "import", None,
                apres={"filename": file.filename, "size": len(content)}, user=identity["email"])
    try:
        report = await _validate_workbook(file.filename or "upload.xlsx", content, identity["email"])
    except Exception as e:
        await audit("import_dryrun_failed", "import", None, apres={"error": "processing_error"}, user=identity["email"])
        raise HTTPException(status_code=400, detail="Échec du traitement du classeur.")
    finally:
        del content  # temp content discarded; nothing persisted to disk

    # Persist ONLY job metadata + validation results (no business data)
    store = {**report, "_summary_rows": len(report["rows"])}
    await mdb.import_jobs.insert_one(dict(store))
    await audit("import_dryrun_complete", "import", report["import_id"],
                apres={"verdict": report["verdict"], "totals": report["totals"]}, user=identity["email"])
    report.pop("_id", None)
    return report


@imports_router.get("")
async def list_imports(identity: dict = Depends(require_role(*IMPORT_ROLES))):
    jobs = await mdb.import_jobs.find({}, {"_id": 0, "rows": 0}).sort("processed_at", -1).to_list(200)
    return jobs


@imports_router.get("/{iid}")
async def get_import(iid: str, identity: dict = Depends(require_role(*IMPORT_ROLES))):
    job = await mdb.import_jobs.find_one({"import_id": iid}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Import introuvable")
    await audit("import_report_access", "import", iid, user=identity["email"])
    return job


@imports_router.delete("/{iid}")
async def delete_import(iid: str, identity: dict = Depends(require_role(*IMPORT_ROLES))):
    job = await mdb.import_jobs.find_one({"import_id": iid}, {"_id": 0, "import_id": 1})
    if not job:
        raise HTTPException(status_code=404, detail="Import introuvable")
    await mdb.import_jobs.delete_one({"import_id": iid})
    await audit("import_cleanup", "import", iid, user=identity["email"])
    return {"deleted": iid}
